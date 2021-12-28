#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import csv
import html
import json
import logging
import multiprocessing as mp
import random
import string
import time
from copy import deepcopy
from io import BytesIO, StringIO
from operator import itemgetter

import cherrypy
from cherrypy import _cperror

from mako.lookup import TemplateLookup
from mako.template import Template

import openpyxl

import secure

from ghostosintlib import GhostOSINT

from ghostosintscan import GhostOsintScan

from ghostosint import GhostOsintDB
from ghostosint import GhostOsintHelp
from ghostosint import __version__
from ghostosint.logger import logListenerSetup, logWorkerSetup

mp.set_start_method("spawn", force=True)


class GhostOsintWEB:
    """GhostOSINT web interface."""

    lookup = TemplateLookup(directories=[''])
    defaultConfig = dict()
    config = dict()
    token = None
    docroot = ''

    def __init__(self: 'GhostOsintWEB', web_config: dict, config: dict, loggingQueue: 'logging.handlers.QueueListener' = None) -> None:
        """Initialize web server.

        Args:
            web_config (dict): config settings for web interface (interface, port, root path)
            config (dict): GhostOSINT config
            loggingQueue: TBD

        Raises:
            TypeError: arg type is invalid
            ValueError: arg value is invalid
        """
        if not isinstance(config, dict):
            raise TypeError(f"config is {type(config)}; expected dict()")
        if not config:
            raise ValueError("config is empty")

        if not isinstance(web_config, dict):
            raise TypeError(f"web_config is {type(web_config)}; expected dict()")
        if not config:
            raise ValueError("web_config is empty")

        self.docroot = web_config.get('root', '/').rstrip('/')

        # 'config' supplied will be the defaults, let's supplement them
        # now with any configuration which may have previously been saved.
        self.defaultConfig = deepcopy(config)
        dbh = GhostOsintDB(self.defaultConfig, init=True)
        GhostOsint = GhostOSINT(self.defaultConfig)
        self.config = GhostOsint.configUnserialize(dbh.configGet(), self.defaultConfig)

        # Set up logging
        if loggingQueue is None:
            self.loggingQueue = mp.Queue()
            logListenerSetup(self.loggingQueue, self.config)
        else:
            self.loggingQueue = loggingQueue
        logWorkerSetup(self.loggingQueue)
        self.log = logging.getLogger(f"ghostosint.{__name__}")

        cherrypy.config.update({
            'error_page.401': self.error_page_401,
            'error_page.404': self.error_page_404,
            'request.error_response': self.error_page
        })

        csp = (
            secure.ContentSecurityPolicy()
            .default_src("'self'")
            .script_src("'self'", "'unsafe-inline'", "blob:")
            .style_src("'self'", "'unsafe-inline'")
            .base_uri("'self'")
            .connect_src("'self'", "data:")
            .frame_src("'self'", 'data:')
            .img_src("'self'", "data:")
        )

        secure_headers = secure.Secure(
            server=secure.Server().set("server"),
            cache=secure.CacheControl().must_revalidate(),
            csp=csp,
            referrer=secure.ReferrerPolicy().no_referrer(),
        )

        cherrypy.config.update({
            "tools.response_headers.on": True,
            "tools.response_headers.headers": secure_headers.framework.cherrypy()
        })

    def error_page(self: 'GhostOsintWEB') -> None:
        """Error page."""
        cherrypy.response.status = 500

        if self.config['_debug']:
            cherrypy.response.body = _cperror.get_error_page(status=500, traceback=_cperror.format_exc())
        else:
            cherrypy.response.body = b"<html><body>Error</body></html>"

    def error_page_401(self: 'GhostOsintWEB', status: str, message: str, traceback: str, version: str) -> str:
        """Unauthorized access HTTP 401 error page.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTML response
        """
        return ""

    def error_page_404(self: 'GhostOsintWEB', status: str, message: str, traceback: str, version: str) -> str:
        """Not found error page 404.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message
            traceback (str): Error stack trace
            version (str): CherryPy version

        Returns:
            str: HTTP response template
        """
        templ = Template(filename='ghostosint/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message='Not Found', docroot=self.docroot, status=status, version=__version__)

    def jsonify_error(self: 'GhostOsintWEB', status: str, message: str) -> dict:
        """Jsonify error response.

        Args:
            status (str): HTTP response status code and message
            message (str): Error message

        Returns:
            dict: HTTP error response template
        """
        cherrypy.response.headers['Content-Type'] = 'application/json'
        cherrypy.response.status = status
        return {
            'error': {
                'http_status': status,
                'message': message,
            }
        }

    def error(self: 'GhostOsintWEB', message: str) -> None:
        """Show generic error page with error message.

        Args:
            message (str): error message

        Returns:
            None
        """
        templ = Template(filename='ghostosint/templates/error.tmpl', lookup=self.lookup)
        return templ.render(message=message, docroot=self.docroot, version=__version__)

    def cleanUserInput(self: 'GhostOsintWEB', inputList: list) -> list:
        """Sanitize user input, poorly.

        Args:
            inputList (list): TBD

        Returns:
            list: sanitized input

        Raises:
            TypeError: inputList type was invalid
        """
        if not isinstance(inputList, list):
            raise TypeError(f"inputList is {type(inputList)}; expected list()")

        ret = list()

        for item in inputList:
            if not item:
                ret.append('')
                continue
            c = html.escape(item, True)
            c = c.replace("'", '&quot;')
            # We don't actually want & translated to &amp;
            c = c.replace("&amp;", "&").replace("&quot;", "\"")
            ret.append(c)

        return ret

    def searchBase(self: 'GhostOsintWEB', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search.

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD

        Returns:
            list: search results
        """
        retdata = []

        if not id and not eventType and not value:
            return retdata

        if not value:
            value = ''

        regex = ""
        if value.startswith("/") and value.endswith("/"):
            regex = value[1:len(value) - 1]
            value = ""

        value = value.replace('*', '%')
        if value in [None, ""] and regex in [None, ""]:
            value = "%"
            regex = ""

        dbh = GhostOsintDB(self.config)
        criteria = {
            'scan_id': id or '',
            'type': eventType or '',
            'value': value or '',
            'regex': regex or '',
        }

        try:
            data = dbh.search(criteria)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            escapeddata = html.escape(row[1])
            escapedsrc = html.escape(row[2])
            retdata.append([lastseen, escapeddata, escapedsrc,
                            row[3], row[5], row[6], row[7], row[8], row[10],
                            row[11], row[4], row[13], row[14]])

        return retdata

    def buildExcel(self: 'GhostOsintWEB', data: list, columnNames: list, sheetNameIndex: int = 0) -> str:
        """Convert supplied raw data into GEXF (Graph Exchange XML Format) format (e.g. for Gephi).

        Args:
            data (list): Scan result as list
            columnNames (list): column names
            sheetNameIndex (int): TBD

        Returns:
            str: Excel workbook
        """
        rowNums = dict()
        workbook = openpyxl.Workbook()
        defaultSheet = workbook.active
        columnNames.pop(sheetNameIndex)
        allowed_sheet_chars = string.ascii_uppercase + string.digits + '_'
        for row in data:
            sheetName = "".join([c for c in str(row.pop(sheetNameIndex)) if c.upper() in allowed_sheet_chars])
            try:
                sheet = workbook[sheetName]
            except KeyError:
                # Create sheet
                workbook.create_sheet(sheetName)
                sheet = workbook[sheetName]
                # Write headers
                for col_num, column_title in enumerate(columnNames, 1):
                    cell = sheet.cell(row=1, column=col_num)
                    cell.value = column_title
                rowNums[sheetName] = 2

            # Write row
            for col_num, cell_value in enumerate(row, 1):
                cell = sheet.cell(row=rowNums[sheetName], column=col_num)
                cell.value = cell_value

            rowNums[sheetName] += 1

        if rowNums:
            workbook.remove(defaultSheet)

        # Sort sheets alphabetically
        workbook._sheets.sort(key=lambda ws: ws.title)

        # Save workbook
        with BytesIO() as f:
            workbook.save(f)
            f.seek(0)
            return f.read()

    #
    # USER INTERFACE PAGES
    #

    @cherrypy.expose
    def scanexportlogs(self: 'GhostOsintWEB', id: str, dialect: str = "excel") -> bytes:
        """Get scan log

        Args:
            id (str): scan ID
            dialect (str): CSV dialect (default: excel)

        Returns:
            bytes: scan logs in CSV format
        """
        dbh = GhostOsintDB(self.config)

        try:
            data = dbh.scanLogs(id, None, None, True)
        except Exception:
            return self.error("Scan ID not found.")

        if not data:
            return self.error("Scan ID not found.")

        fileobj = StringIO()
        parser = csv.writer(fileobj, dialect=dialect)
        parser.writerow(["数据", "组成t", "类型", "事件", "事件ID"])
        for row in data:
            parser.writerow([
                time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000)),
                str(row[1]),
                str(row[2]),
                str(row[3]),
                row[4]
            ])

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename=GhostOSINT-{id}.log.csv"
        cherrypy.response.headers['Content-Type'] = "application/csv"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return fileobj.getvalue().encode('utf-8')

    @cherrypy.expose
    def scaneventresultexport(self: 'GhostOsintWEB', id: str, type: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format

        Args:
            id (str): scan ID
            type (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = GhostOsintDB(self.config)
        data = dbh.scanResultEvent(id, type)

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "GhostOSINT.xlsx"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["已更新", "类型", "模块", "源",
                                   "内容", "数据"], sheetNameIndex=1)

        elif filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["已更新", "类型", "模块", "源", "内容", "数据"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([lastseen, str(row[4]), str(row[3]), str(row[2]), row[13], datafield])

            fname = "GhostOSINT.csv"
            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        else:
            return self.error("导出文件类型失败.")

    @cherrypy.expose
    def scaneventresultexportmulti(self: 'GhostOsintWEB', ids: str, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get scan event result data in CSV or Excel format for multiple scans

        Args:
            ids (str): comma separated list of scan IDs
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        dbh = GhostOsintDB(self.config)
        scaninfo = dict()
        data = list()
        scan_name = ""

        for id in ids.split(','):
            scaninfo[id] = dbh.scanInstanceGet(id)
            if scaninfo[id] is None:
                continue
            scan_name = scaninfo[id][0]
            data = data + dbh.scanResultEvent(id)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                            str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "GhostOSINT.xlsx"
            else:
                fname = scan_name + "-GhostOSINT.xlsx"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["扫描名称", "已更新", "类型", "模块",
                                   "源", "内容", "数据"], sheetNameIndex=2)

        elif filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["扫描名称", "已更新", "类型", "模块", "源", "内容", "数据"])
            for row in data:
                if row[4] == "ROOT":
                    continue
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([scaninfo[row[12]][0], lastseen, str(row[4]), str(row[3]),
                                str(row[2]), row[13], datafield])

            if len(ids.split(',')) > 1 or scan_name == "":
                fname = "GhostOSINT.csv"
            else:
                fname = scan_name + "-GhostOSINT.csv"

            cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        else:
            return self.error("导出文件类型无效.")

    @cherrypy.expose
    def scansearchresultexport(self: 'GhostOsintWEB', id: str, eventType: str = None, value: str = None, filetype: str = "csv", dialect: str = "excel") -> str:
        """Get search result data in CSV or Excel format

        Args:
            id (str): scan ID
            eventType (str): TBD
            value (str): TBD
            filetype (str): type of file ("xlsx|excel" or "csv")
            dialect (str): CSV dialect (default: excel)

        Returns:
            str: results in CSV or Excel format
        """
        data = self.searchBase(id, eventType, value)

        if not data:
            return None

        if filetype.lower() in ["xlsx", "excel"]:
            rows = []
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                rows.append([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=GhostOSINT.xlsx"
            cherrypy.response.headers['Content-Type'] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return self.buildExcel(rows, ["已更新", "类型", "模块", "源",
                                   "内容", "数据"], sheetNameIndex=1)

        elif filetype.lower() == 'csv':
            fileobj = StringIO()
            parser = csv.writer(fileobj, dialect=dialect)
            parser.writerow(["已更新", "类型", "模块", "源", "内容", "数据"])
            for row in data:
                if row[10] == "ROOT":
                    continue
                datafield = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                parser.writerow([row[0], str(row[10]), str(row[3]), str(row[2]), row[11], datafield])
            cherrypy.response.headers['Content-Disposition'] = "attachment; filename=GhostOSINT.csv"
            cherrypy.response.headers['Content-Type'] = "application/csv"
            cherrypy.response.headers['Pragma'] = "no-cache"
            return fileobj.getvalue().encode('utf-8')

        else:
            return self.error("Invalid export filetype.")

    @cherrypy.expose
    def scanexportjsonmulti(self: 'GhostOsintWEB', ids: str) -> str:
        """Get scan event result data in JSON format for multiple scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: results in JSON format
        """
        dbh = GhostOsintDB(self.config)
        scaninfo = list()
        scan_name = ""

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)

            if scan is None:
                continue

            scan_name = scan[0]

            for row in dbh.scanResultEvent(id):
                lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
                event_data = str(row[1]).replace("<SFURL>", "").replace("</SFURL>", "")
                source_data = str(row[2])
                source_module = str(row[3])
                event_type = row[4]
                false_positive = row[13]

                if event_type == "ROOT":
                    continue

                scaninfo.append({
                    "data": event_data,
                    "event_type": event_type,
                    "module": source_module,
                    "source_data": source_data,
                    "false_positive": false_positive,
                    "last_seen": lastseen,
                    "scan_name": scan_name,
                    "scan_target": scan[1]
                })

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "GhostOSINT.json"
        else:
            fname = scan_name + "-GhostOSINT.json"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return json.dumps(scaninfo).encode('utf-8')

    @cherrypy.expose
    def scanviz(self: 'GhostOsintWEB', id: str, gexf: str = "0") -> str:
        """Export entities from scan results for visualising.

        Args:
            id (str): scan ID
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        if not id:
            return None

        dbh = GhostOsintDB(self.config)
        data = dbh.scanResultEvent(id, filterFp=True)
        scan = dbh.scanInstanceGet(id)

        if not scan:
            return None

        scan_name = scan[0]

        root = scan[1]

        if gexf == "0":
            return GhostOsintHelp.buildGraphJson([root], data)

        if not scan_name:
            fname = "GhostOSINT.gexf"
        else:
            fname = scan_name + "GhostOSINT.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return GhostOsintHelp.buildGraphGexf([root], "GhostOSINT Export", data)

    @cherrypy.expose
    def scanvizmulti(self: 'GhostOsintWEB', ids: str, gexf: str = "1") -> str:
        """Export entities results from multiple scans in GEXF format.

        Args:
            ids (str): scan IDs
            gexf (str): TBD

        Returns:
            str: GEXF data
        """
        dbh = GhostOsintDB(self.config)
        data = list()
        roots = list()
        scan_name = ""

        if not ids:
            return None

        for id in ids.split(','):
            scan = dbh.scanInstanceGet(id)
            if not scan:
                continue
            data = data + dbh.scanResultEvent(id, filterFp=True)
            roots.append(scan[1])
            scan_name = scan[0]

        if not data:
            return None

        if gexf == "0":
            # Not implemented yet
            return None

        if len(ids.split(',')) > 1 or scan_name == "":
            fname = "GhostOSINT.gexf"
        else:
            fname = scan_name + "-GhostOSINT.gexf"

        cherrypy.response.headers['Content-Disposition'] = f"attachment; filename={fname}"
        cherrypy.response.headers['Content-Type'] = "application/gexf"
        cherrypy.response.headers['Pragma'] = "no-cache"
        return GhostOsintHelp.buildGraphGexf(roots, "GhostOSINT Export", data)

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanopts(self: 'GhostOsintWEB', id: str) -> str:
        """Return configuration used for the specified scan as JSON.

        Args:
            id: scan ID

        Returns:
            str: options as JSON string
        """
        dbh = GhostOsintDB(self.config)
        ret = dict()

        meta = dbh.scanInstanceGet(id)
        if not meta:
            return ret

        if meta[3] != 0:
            started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[3]))
        else:
            started = "Not yet"

        if meta[4] != 0:
            finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(meta[4]))
        else:
            finished = "Not yet"

        ret['meta'] = [meta[0], meta[1], meta[2], started, finished, meta[5]]
        ret['config'] = dbh.scanConfigGet(id)
        ret['configdesc'] = dict()
        for key in list(ret['config'].keys()):
            if ':' not in key:
                globaloptdescs = self.config['__globaloptdescs__']
                if globaloptdescs:
                    ret['configdesc'][key] = globaloptdescs.get(key, f"{key} (legacy)")
            else:
                [modName, modOpt] = key.split(':')
                if modName not in list(self.config['__modules__'].keys()):
                    continue

                if modOpt not in list(self.config['__modules__'][modName]['optdescs'].keys()):
                    continue

                ret['configdesc'][key] = self.config['__modules__'][modName]['optdescs'][modOpt]

        return ret

    @cherrypy.expose
    def rerunscan(self: 'GhostOsintWEB', id: str) -> None:
        """Rerun a scan.

        Args:
            id (str): scan ID

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to info page for new scan
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = GhostOsintDB(cfg)
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("无效扫描ID.")

        scanname = info[0]
        scantarget = info[1]

        scanconfig = dbh.scanConfigGet(id)
        if not scanconfig:
            return self.error(f"扫描加载配置文件出错: {id}")

        modlist = scanconfig['_modulesenabled'].split(',')
        if "GO__stor_stdout" in modlist:
            modlist.remove("GO__stor_stdout")

        targetType = GhostOsintHelp.targetTypeFromString(scantarget)
        if not targetType:
            # It must then be a name, as a re-run scan should always have a clean
            # target. Put quotes around the target value and try to determine the
            # target type again.
            targetType = GhostOsintHelp.targetTypeFromString(f'"{scantarget}"')

        if targetType not in ["HUMAN_NAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = GhostOsintHelp.genScanInstanceId()
        try:
            p = mp.Process(target=GhostOsintScan, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] 扫描 [{scanId}] 失败: {e}")
            return self.error(f"[-] 扫描 [{scanId}] 失败: {e}")

        # Wait until the scan has initialized
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("正在等待扫描初始化...")
            time.sleep(1)

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}", status=302)

    @cherrypy.expose
    def rerunscanmulti(self: 'GhostOsintWEB', ids: str) -> str:
        """Rerun scans.

        Args:
            ids (str): comma separated list of scan IDs

        Returns:
            str: Scan list page HTML
        """
        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        modlist = list()
        dbh = GhostOsintDB(cfg)

        for id in ids.split(","):
            info = dbh.scanInstanceGet(id)
            if not info:
                return self.error("无效扫描ID.")

            scanconfig = dbh.scanConfigGet(id)
            scanname = info[0]
            scantarget = info[1]
            targetType = None

            if len(scanconfig) == 0:
                return self.error("Something went wrong internally.")

            modlist = scanconfig['_modulesenabled'].split(',')
            if "GO__stor_stdout" in modlist:
                modlist.remove("GO__stor_stdout")

            targetType = GhostOsintHelp.targetTypeFromString(scantarget)
            if targetType is None:
                # Should never be triggered for a re-run scan..
                return self.error("无效目标类型. 无法识别目标为 GhostOSINT 支持的类型.")

            # Start running a new scan
            scanId = GhostOsintHelp.genScanInstanceId()
            try:
                p = mp.Process(target=GhostOsintScan, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
                p.daemon = True
                p.start()
            except Exception as e:
                self.log.error(f"[-] 扫描 [{scanId}] 失败: {e}")
                return self.error(f"[-] 扫描 [{scanId}] 失败: {e}")

            # Wait until the scan has initialized
            while dbh.scanInstanceGet(scanId) is None:
                self.log.info("正在等待扫描初始化...")
                time.sleep(1)

        templ = Template(filename='ghostosint/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(rerunscans=True, docroot=self.docroot, pageid="SCANLIST", version=__version__)

    @cherrypy.expose
    def newscan(self: 'GhostOsintWEB') -> str:
        """Configure a new scan.

        Returns:
            str: New scan page HTML
        """
        dbh = GhostOsintDB(self.config)
        types = dbh.eventTypes()
        templ = Template(filename='ghostosint/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], scanname="",
                            selectedmods="", scantarget="", version=__version__)

    @cherrypy.expose
    def clonescan(self: 'GhostOsintWEB', id: str) -> str:
        """Clone an existing scan (pre-selected options in the newscan page).

        Args:
            id (str): scan ID to clone

        Returns:
            str: New scan page HTML pre-populated with options from cloned scan.
        """
        dbh = GhostOsintDB(self.config)
        types = dbh.eventTypes()
        info = dbh.scanInstanceGet(id)

        if not info:
            return self.error("无效扫描ID.")

        scanconfig = dbh.scanConfigGet(id)
        scanname = info[0]
        scantarget = info[1]
        targetType = None

        if scanname == "" or scantarget == "" or len(scanconfig) == 0:
            return self.error("Something went wrong internally.")

        targetType = GhostOsintHelp.targetTypeFromString(scantarget)
        if targetType is None:
            # It must be a name, so wrap quotes around it
            scantarget = "&quot;" + scantarget + "&quot;"

        modlist = scanconfig['_modulesenabled'].split(',')

        templ = Template(filename='ghostosint/templates/newscan.tmpl', lookup=self.lookup)
        return templ.render(pageid='NEWSCAN', types=types, docroot=self.docroot,
                            modules=self.config['__modules__'], selectedmods=modlist,
                            scanname=str(scanname),
                            scantarget=str(scantarget), version=__version__)

    @cherrypy.expose
    def index(self: 'GhostOsintWEB') -> str:
        """Show scan list page.

        Returns:
            str: Scan list page HTML
        """
        templ = Template(filename='ghostosint/templates/scanlist.tmpl', lookup=self.lookup)
        return templ.render(pageid='SCANLIST', docroot=self.docroot, version=__version__)

    @cherrypy.expose
    def scaninfo(self: 'GhostOsintWEB', id: str) -> str:
        """Information about a selected scan.

        Args:
            id (str): scan id

        Returns:
            str: scan info page HTML
        """
        dbh = GhostOsintDB(self.config)
        res = dbh.scanInstanceGet(id)
        if res is None:
            return self.error("Scan ID not found.")

        templ = Template(filename='ghostosint/templates/scaninfo.tmpl', lookup=self.lookup, input_encoding='utf-8')
        return templ.render(id=id, name=html.escape(res[0]), status=res[5], docroot=self.docroot, version=__version__,
                            pageid="SCANLIST")

    @cherrypy.expose
    def opts(self: 'GhostOsintWEB', updated: str = None) -> str:
        """Show module and global settings page.

        Args:
            updated (str): scan options were updated successfully

        Returns:
            str: scan options page HTML
        """
        templ = Template(filename='ghostosint/templates/opts.tmpl', lookup=self.lookup)
        self.token = random.SystemRandom().randint(0, 99999999)
        return templ.render(opts=self.config, pageid='SETTINGS', token=self.token, version=__version__,
                            updated=updated, docroot=self.docroot)

    @cherrypy.expose
    def optsexport(self: 'GhostOsintWEB', pattern: str = None) -> str:
        """Export configuration.

        Args:
            pattern (str): TBD

        Returns:
            str: Configuration settings
        """
        GhostOsint = GhostOSINT(self.config)
        conf = GhostOsint.configSerialize(self.config)
        content = ""

        for opt in sorted(conf):
            if ":_" in opt or opt.startswith("_"):
                continue

            if pattern:
                if pattern in opt:
                    content += f"{opt}={conf[opt]}\n"
            else:
                content += f"{opt}={conf[opt]}\n"

        cherrypy.response.headers['Content-Disposition'] = 'attachment; filename="GhostOSINT.cfg"'
        cherrypy.response.headers['Content-Type'] = "text/plain"
        return content

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def optsraw(self: 'GhostOsintWEB') -> str:
        """Return global and module settings as json.

        Returns:
            str: settings as JSON
        """
        ret = dict()
        self.token = random.SystemRandom().randint(0, 99999999)
        for opt in self.config:
            if not opt.startswith('__'):
                ret["global." + opt] = self.config[opt]
                continue

            if opt == '__modules__':
                for mod in sorted(self.config['__modules__'].keys()):
                    for mo in sorted(self.config['__modules__'][mod]['opts'].keys()):
                        if mo.startswith("_"):
                            continue
                        ret["module." + mod + "." + mo] = self.config['__modules__'][mod]['opts'][mo]

        return ['SUCCESS', {'token': self.token, 'data': ret}]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scandelete(self: 'GhostOsintWEB', id: str) -> str:
        """Delete scan(s).

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "未指定扫描")

        dbh = GhostOsintDB(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"扫描 {scan_id} 不存在")

            if res[5] in ["RUNNING", "STARTING", "STARTED"]:
                return self.jsonify_error('400', f"扫描 {scan_id} 是 {res[5]}. 你不能删除正在运行的扫描.")

        for scan_id in ids:
            dbh.scanInstanceDelete(scan_id)

        return ""

    @cherrypy.expose
    def savesettings(self: 'GhostOsintWEB', allopts: str, token: str, configFile: 'cherrypy._cpreqbody.Part' = None) -> None:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token
            configFile (cherrypy._cpreqbody.Part): TBD

        Returns:
            None

        Raises:
            HTTPRedirect: redirect to scan settings
        """
        if str(token) != str(self.token):
            return self.error(f"无效Token ({token})")

        if configFile:  # configFile seems to get set even if a file isn't uploaded
            if configFile.file:
                contents = configFile.file.read()

                if isinstance(contents, bytes):
                    contents = contents.decode('utf-8')

                try:
                    tmp = dict()
                    for line in contents.split("\n"):
                        if "=" not in line:
                            continue

                        opt_array = line.strip().split("=")
                        if len(opt_array) == 1:
                            opt_array[1] = ""

                        tmp[opt_array[0]] = '='.join(opt_array[1:])

                    allopts = json.dumps(tmp).encode('utf-8')
                except Exception as e:
                    return self.error(f"无法分析输入的文件. 文件是 Ghost OSINT生成的吗? ({e})")

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")
            return self.error("无法重置设置")

        # Save settings
        try:
            dbh = GhostOsintDB(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            GhostOsint = GhostOSINT(self.config)
            self.config = GhostOsint.configUnserialize(cleanopts, currentopts)
            dbh.configSet(GhostOsint.configSerialize(self.config))
        except Exception as e:
            return self.error(f"处理一个或多个输入失败: {e}")

        raise cherrypy.HTTPRedirect(f"{self.docroot}/opts?updated=1")

    @cherrypy.expose
    def savesettingsraw(self: 'GhostOsintWEB', allopts: str, token: str) -> str:
        """Save settings, also used to completely reset them to default.

        Args:
            allopts: TBD
            token (str): CSRF token

        Returns:
            str: save success as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        if str(token) != str(self.token):
            return json.dumps(["ERROR", f"无效 Token ({token})."]).encode('utf-8')

        # Reset config to default
        if allopts == "RESET":
            if self.reset_settings():
                return json.dumps(["SUCCESS", ""]).encode('utf-8')
            return json.dumps(["ERROR", "无法重置设置"]).encode('utf-8')

        # Save settings
        try:
            dbh = GhostOsintDB(self.config)
            useropts = json.loads(allopts)
            cleanopts = dict()
            for opt in list(useropts.keys()):
                cleanopts[opt] = self.cleanUserInput([useropts[opt]])[0]

            currentopts = deepcopy(self.config)

            # Make a new config where the user options override
            # the current system config.
            GhostOsint = GhostOSINT(self.config)
            self.config = GhostOsint.configUnserialize(cleanopts, currentopts)
            dbh.configSet(GhostOsint.configSerialize(self.config))
        except Exception as e:
            return json.dumps(["ERROR", f"处理一个或多个输入失败: {e}"]).encode('utf-8')

        return json.dumps(["SUCCESS", ""]).encode('utf-8')

    def reset_settings(self: 'GhostOsintWEB') -> bool:
        """Reset settings to default.

        Returns:
            bool: success
        """
        try:
            dbh = GhostOsintDB(self.config)
            dbh.configClear()  # Clear it in the DB
            self.config = deepcopy(self.defaultConfig)  # Clear in memory
        except Exception:
            return False

        return True

    @cherrypy.expose
    def resultsetfp(self: 'GhostOsintWEB', id: str, resultids: str, fp: str) -> str:
        """Set a bunch of results (hashes) as false positive.

        Args:
            id (str): scan ID
            resultids (str): comma separated list of result IDs
            fp (str): 0 or 1

        Returns:
            str: set false positive status as JSON
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = GhostOsintDB(self.config)

        if fp not in ["0", "1"]:
            return json.dumps(["ERROR", "未设置误报标志或设置不正确."]).encode('utf-8')

        try:
            ids = json.loads(resultids)
        except Exception:
            return json.dumps(["ERROR", "未提供ID."]).encode('utf-8')

        # Cannot set FPs if a scan is not completed
        status = dbh.scanInstanceGet(id)
        if not status:
            return self.error(f"无效扫描ID: {id}")

        if status[5] not in ["ABORTED", "FINISHED", "ERROR-FAILED"]:
            return json.dumps([
                "WARNING",
                "Scan must be in a finished state when setting False Positives."
            ]).encode('utf-8')

        # Make sure the user doesn't set something as non-FP when the
        # parent is set as an FP.
        if fp == "0":
            data = dbh.scanElementSourcesDirect(id, ids)
            for row in data:
                if str(row[14]) == "1":
                    return json.dumps([
                        "WARNING",
                        f"Cannot unset element {id} as False Positive if a parent element is still False Positive."
                    ]).encode('utf-8')

        # Set all the children as FPs too.. it's only logical afterall, right?
        childs = dbh.scanElementChildrenAll(id, ids)
        allIds = ids + childs

        ret = dbh.scanResultsUpdateFP(id, allIds, fp)
        if ret:
            return json.dumps(["SUCCESS", ""]).encode('utf-8')

        return json.dumps(["ERROR", "遇到异常."]).encode('utf-8')

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def eventtypes(self: 'GhostOsintWEB') -> str:
        """List all event types.

        Returns:
            str: list of event types
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        dbh = GhostOsintDB(self.config)
        types = dbh.eventTypes()
        ret = list()

        for r in types:
            ret.append([r[1], r[0]])

        return sorted(ret, key=itemgetter(0))

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def modules(self: 'GhostOsintWEB') -> str:
        """List all modules.

        Returns:
            str: list of modules
        """
        cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"

        modinfo = list(self.config['__modules__'].keys())
        modinfo.sort()
        ret = list()
        for m in modinfo:
            if "__" in m:
                continue
            ret.append({'name': m, 'descr': self.config['__modules__'][m]['descr']})

        return ret

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def ping(self: 'GhostOsintWEB') -> list:
        """For the CLI to test connectivity to this server.

        Returns:
            list: GhostOSINT version as JSON
        """
        return ["SUCCESS", __version__]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def query(self: 'GhostOsintWEB', query: str) -> str:
        """For the CLI to run queries against the database.

        Args:
            query (str): SQL query

        Returns:
            str: query results as JSON
        """
        dbh = GhostOsintDB(self.config)

        if not query:
            return self.jsonify_error('400', "查询无效.")

        if not query.lower().startswith("select"):
            return self.jsonify_error('400', "不选择也可能会出错滴.")

        try:
            ret = dbh.dbh.execute(query)
            data = ret.fetchall()
            columnNames = [c[0] for c in dbh.dbh.description]
            return [dict(zip(columnNames, row)) for row in data]
        except Exception as e:
            return self.jsonify_error('500', str(e))

    @cherrypy.expose
    def startscan(self: 'GhostOsintWEB', scanname: str, scantarget: str, modulelist: str, typelist: str, usecase: str) -> str:
        """Initiate a scan.

        Args:
            scanname (str): scan name
            scantarget (str): scan target
            modulelist (str): comma separated list of modules to use
            typelist (str): selected modules based on produced event data types
            usecase (str): selected module group (passive, investigate, footprint, all)

        Returns:
            str: start scan status as JSON

        Raises:
            HTTPRedirect: redirect to new scan info page
        """
        [scanname, scantarget] = self.cleanUserInput([scanname, scantarget])

        if not scanname:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "用法不正确: 未指定扫描名称."]).encode('utf-8')

            return self.error("请求无效: 未指定扫描名称.")

        if not scantarget:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "用法不正确: 未指定扫描目标."]).encode('utf-8')

            return self.error("请求无效: 未指定扫描目标.")

        if not typelist and not modulelist and not usecase:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "用法不正确: 没有为扫描指定模块."]).encode('utf-8')

            return self.error("请求无效: 未指定扫描模块.")

        targetType = GhostOsintHelp.targetTypeFromString(scantarget)
        if targetType is None:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "无法识别的目标类型."]).encode('utf-8')

            return self.error("目标类型无效. 无法识别为其是否是 Ghost OSINT 支持的目标类型.")

        # Swap the globalscantable for the database handler
        dbh = GhostOsintDB(self.config)

        # Snapshot the current configuration to be used by the scan
        cfg = deepcopy(self.config)
        GhostOsint = GhostOSINT(cfg)

        modlist = list()

        # User selected modules
        if modulelist:
            modlist = modulelist.replace('module_', '').split(',')

        # User selected types
        if len(modlist) == 0 and typelist:
            typesx = typelist.replace('type_', '').split(',')

            # 1. Find all modules that produce the requested types
            modlist = GhostOsint.modulesProducing(typesx)
            newmods = deepcopy(modlist)
            newmodcpy = deepcopy(newmods)

            # 2. For each type those modules consume, get modules producing
            while len(newmodcpy) > 0:
                for etype in GhostOsint.eventsToModules(newmodcpy):
                    xmods = GhostOsint.modulesProducing([etype])
                    for mod in xmods:
                        if mod not in modlist:
                            modlist.append(mod)
                            newmods.append(mod)
                newmodcpy = deepcopy(newmods)
                newmods = list()

        # User selected a use case
        if len(modlist) == 0 and usecase:
            for mod in self.config['__modules__']:
                if usecase == 'all' or usecase in self.config['__modules__'][mod]['group']:
                    modlist.append(mod)

        # If we somehow got all the way through to here and still don't have any modules selected
        if not modlist:
            if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
                cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
                return json.dumps(["ERROR", "用法不正确: 没有为扫描指定模块."]).encode('utf-8')

            return self.error("请求无效: 未指定扫描模块.")

        # Add our mandatory storage module
        if "GO__stor_db" not in modlist:
            modlist.append("GO__stor_db")
        modlist.sort()

        # Delete the stdout module in case it crept in
        if "GO__stor_stdout" in modlist:
            modlist.remove("GO__stor_stdout")

        # Start running a new scan
        if targetType in ["HUMAN_NAME", "USERNAME", "BITCOIN_ADDRESS"]:
            scantarget = scantarget.replace("\"", "")
        else:
            scantarget = scantarget.lower()

        # Start running a new scan
        scanId = GhostOsintHelp.genScanInstanceId()
        try:
            p = mp.Process(target=GhostOsintScan, args=(self.loggingQueue, scanname, scanId, scantarget, targetType, modlist, cfg))
            p.daemon = True
            p.start()
        except Exception as e:
            self.log.error(f"[-] 扫描 [{scanId}] 失败了: {e}")
            return self.error(f"[-] 扫描 [{scanId}] 失败了: {e}")

        # Wait until the scan has initialized
        # Check the database for the scan status results
        while dbh.scanInstanceGet(scanId) is None:
            self.log.info("正在等待扫描初始化...")
            time.sleep(1)

        if cherrypy.request.headers.get('Accept') and 'application/json' in cherrypy.request.headers.get('Accept'):
            cherrypy.response.headers['Content-Type'] = "application/json; charset=utf-8"
            return json.dumps(["SUCCESS", scanId]).encode('utf-8')

        raise cherrypy.HTTPRedirect(f"{self.docroot}/scaninfo?id={scanId}")

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def stopscan(self: 'GhostOsintWEB', id: str) -> str:
        """Stop a scan.

        Args:
            id (str): comma separated list of scan IDs

        Returns:
            str: JSON response
        """
        if not id:
            return self.jsonify_error('404', "未指定扫描")

        dbh = GhostOsintDB(self.config)
        ids = id.split(',')

        for scan_id in ids:
            res = dbh.scanInstanceGet(scan_id)
            if not res:
                return self.jsonify_error('404', f"扫描 {scan_id} 不存在")

            scan_status = res[5]

            if scan_status == "FINISHED":
                return self.jsonify_error('400', f"扫描 {scan_id} 已经完成了.")

            if scan_status == "ABORTED":
                return self.jsonify_error('400', f"扫描 {scan_id} 已经中止了.")

            if scan_status != "RUNNING" and scan_status != "STARTING":
                return self.jsonify_error('400', f"正在运行的扫描当前处于 '{scan_status}', 请稍后重试或重新启动 Ghost OSINT.")

        for scan_id in ids:
            dbh.scanInstanceSet(scan_id, status="ABORT-REQUESTED")

        return ""

    #
    # DATA PROVIDERS
    #

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlog(self: 'GhostOsintWEB', id: str, limit: str = None, rowId: str = None, reverse: str = None) -> list:
        """Scan log data.

        Args:
            id (str): scan ID
            limit (str): TBD
            rowId (str): TBD
            reverse (str): TBD

        Returns:
            list: scan log
        """
        dbh = GhostOsintDB(self.config)
        retdata = []

        try:
            data = dbh.scanLogs(id, limit, rowId, reverse)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], row[2], html.escape(row[3]), row[4]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanerrors(self: 'GhostOsintWEB', id: str, limit: str = None) -> list:
        """Scan error data.

        Args:
            id (str): scan ID
            limit (str): limit number of results

        Returns:
            list: scan errors
        """
        dbh = GhostOsintDB(self.config)
        retdata = []

        try:
            data = dbh.scanErrors(id, limit)
        except Exception:
            return retdata

        for row in data:
            generated = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0] / 1000))
            retdata.append([generated, row[1], html.escape(str(row[2]))])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanlist(self: 'GhostOsintWEB') -> list:
        """Produce a list of scans.

        Returns:
            list: scan list
        """
        dbh = GhostOsintDB(self.config)
        data = dbh.scanInstanceList()
        retdata = []

        for row in data:
            created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[3]))

            if row[4] == 0:
                started = "Not yet"
            else:
                started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[4]))

            if row[5] == 0:
                finished = "Not yet"
            else:
                finished = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[5]))

            retdata.append([row[0], row[1], row[2], created, started, finished, row[6], row[7]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanstatus(self: 'GhostOsintWEB', id: str) -> list:
        """Show basic information about a scan, including status and number of each event type.

        Args:
            id (str): scan ID

        Returns:
            list: scan status
        """
        dbh = GhostOsintDB(self.config)
        data = dbh.scanInstanceGet(id)

        if not data:
            return []

        created = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[2]))
        started = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[3]))
        ended = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(data[4]))

        return [data[0], data[1], created, started, ended, data[5]]

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scansummary(self: 'GhostOsintWEB', id: str, by: str) -> list:
        """Summary of scan results.

        Args:
            id (str): scan ID
            by (str): filter by type

        Returns:
            list: scan summary
        """
        retdata = []

        dbh = GhostOsintDB(self.config)

        try:
            scandata = dbh.scanResultSummary(id, by)
        except Exception:
            return retdata

        try:
            statusdata = dbh.scanInstanceGet(id)
        except Exception:
            return retdata

        for row in scandata:
            if row[0] == "ROOT":
                continue
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[2]))
            retdata.append([row[0], row[1], lastseen, row[3], row[4], statusdata[5]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresults(self: 'GhostOsintWEB', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return all event results for a scan as JSON.

        Args:
            id (str): scan ID
            eventType (str): filter by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: scan results
        """
        retdata = []

        dbh = GhostOsintDB(self.config)

        try:
            data = dbh.scanResultEvent(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            lastseen = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(row[0]))
            retdata.append([
                lastseen,
                html.escape(row[1]),
                html.escape(row[2]),
                row[3],
                row[5],
                row[6],
                row[7],
                row[8],
                row[13],
                row[14],
                row[4]
            ])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scaneventresultsunique(self: 'GhostOsintWEB', id: str, eventType: str, filterfp: bool = False) -> list:
        """Return unique event results for a scan as JSON.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            filterfp (bool): remove false positives from search results

        Returns:
            list: unique search results
        """
        dbh = GhostOsintDB(self.config)
        retdata = []

        try:
            data = dbh.scanResultEventUnique(id, eventType, filterfp)
        except Exception:
            return retdata

        for row in data:
            escaped = html.escape(row[0])
            retdata.append([escaped, row[1], row[2]])

        return retdata

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def search(self: 'GhostOsintWEB', id: str = None, eventType: str = None, value: str = None) -> list:
        """Search scans.

        Args:
            id (str): filter search results by scan ID
            eventType (str): filter search results by event type
            value (str): filter search results by event value

        Returns:
            list: search results
        """
        try:
            return self.searchBase(id, eventType, value)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanhistory(self: 'GhostOsintWEB', id: str) -> list:
        """Historical data for a scan.

        Args:
            id (str): scan ID

        Returns:
            list: scan history
        """
        if not id:
            return self.jsonify_error('404', "未指定扫描")

        dbh = GhostOsintDB(self.config)

        try:
            return dbh.scanResultHistory(id)
        except Exception:
            return []

    @cherrypy.expose
    @cherrypy.tools.json_out()
    def scanelementtypediscovery(self: 'GhostOsintWEB', id: str, eventType: str) -> dict:
        """Scan element type discovery.

        Args:
            id (str): scan ID
            eventType (str): filter by event type

        Returns:
            dict
        """
        dbh = GhostOsintDB(self.config)
        pc = dict()
        datamap = dict()
        retdata = dict()

        # Get the events we will be tracing back from
        try:
            leafSet = dbh.scanResultEvent(id, eventType)
            [datamap, pc] = dbh.scanElementSourcesAll(id, leafSet)
        except Exception:
            return retdata

        # Delete the ROOT key as it adds no value from a viz perspective
        del pc['ROOT']
        retdata['tree'] = GhostOsintHelp.dataParentChildToTree(pc)
        retdata['data'] = datamap

        return retdata
